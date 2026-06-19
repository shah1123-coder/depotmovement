"""Open a workbook (decrypt if needed) and return detection + extraction views.

A 'view' here is a uniform wrapper around either an openpyxl worksheet or a
legacy .xls sheet, exposing the same minimal interface the grid builder needs:
  - title
  - max_row, max_col (1-based)
  - cell(row, col) -> value
  - merged_ranges() -> list[(min_row, min_col, max_row, max_col)]
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Optional

from ..errors import Failure, StageResult
from ..logging import get_logger

log = get_logger("workbook.loader")

_XLS_EXTS = {".xls"}
_XLSX_EXTS = {".xlsx", ".xlsm"}


class SheetView:
    def __init__(self, title, max_row, max_col, cell_fn, merged):
        self.title = title
        self.max_row = max_row
        self.max_col = max_col
        self._cell_fn = cell_fn
        self._merged = merged

    def cell(self, row: int, col: int):
        return self._cell_fn(row, col)

    def merged_ranges(self):
        return self._merged


class WorkbookViews:
    def __init__(self, detection: list[SheetView], extraction: list[SheetView]):
        self.detection = detection
        self.extraction = extraction


def _decrypt_if_needed(path: Path) -> io.BytesIO | Path:
    try:
        import msoffcrypto
    except ImportError:
        return path
    try:
        with open(path, "rb") as fh:
            office = msoffcrypto.OfficeFile(fh)
            if not office.is_encrypted():
                return path
            office.load_key(password="")  # common depot convention: blank/known pwd
            out = io.BytesIO()
            office.decrypt(out)
            out.seek(0)
            return out
    except Exception:
        return path


def _openpyxl_views(source) -> tuple[list[SheetView], list[SheetView]]:
    from openpyxl import load_workbook

    def build(data_only: bool) -> list[SheetView]:
        if isinstance(source, io.BytesIO):
            source.seek(0)
        wb = load_workbook(source, data_only=data_only, read_only=False)
        views = []
        for ws in wb.worksheets:
            merged = [(r.min_row, r.min_col, r.max_row, r.max_col)
                      for r in ws.merged_cells.ranges]
            views.append(SheetView(
                ws.title, ws.max_row or 0, ws.max_column or 0,
                (lambda ws: lambda r, c: ws.cell(row=r, column=c).value)(ws),
                merged,
            ))
        return views

    return build(False), build(True)


def _xls_views(path: Path) -> tuple[list[SheetView], list[SheetView]]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)

    def build() -> list[SheetView]:
        views = []
        for sh in book.sheets():
            merged = [(r0 + 1, c0 + 1, r1, c1) for (r0, r1, c0, c1) in sh.merged_cells]

            def make_cell(sh):
                def cell(row, col):
                    if 1 <= row <= sh.nrows and 1 <= col <= sh.ncols:
                        v = sh.cell_value(row - 1, col - 1)
                        return v if v != "" else None
                    return None
                return cell

            views.append(SheetView(sh.name, sh.nrows, sh.ncols, make_cell(sh), merged))
        return views

    # .xls has no formula/value distinction here; both views are identical.
    return build(), build()


def open_workbook(path: str | Path) -> StageResult[WorkbookViews]:
    p = Path(path)
    try:
        ext = p.suffix.lower()
        if ext in _XLS_EXTS:
            det, ext_v = _xls_views(p)
        else:
            source = _decrypt_if_needed(p)
            det, ext_v = _openpyxl_views(source)
        return StageResult.ok(WorkbookViews(det, ext_v))
    except Exception as exc:
        return StageResult.fail(Failure(
            stage="workbook.loader", file=str(p),
            reason="open_failed", detail=str(exc)))
