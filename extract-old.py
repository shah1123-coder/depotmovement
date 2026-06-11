#!/usr/bin/env python3
"""
extract.py

Detects the (possibly multiple) tables inside each sheet of a spreadsheet and
extracts each table — with its title — into its own CSV, retaining structure.

Output layout (under an "extraction" folder next to this script):
    extraction/<file>/<sheet>/<table-title>.csv
A table with no detectable title is named after its sheet. A sheet with no
detectable table falls back to dumping the whole sheet.

Usage:
    python extract.py <file1> [file2 ...]
    python extract.py            # processes every spreadsheet next to this script

Supported inputs: .csv .tsv .txt .xlsx .xlsm .xltx .xltm .xls .xlsb .ods
Anything else is attempted via a pandas fallback if pandas is installed.
"""

import csv
import datetime as _dt
import re
import sys
from pathlib import Path

# ISO 6346 container number, e.g. MSCU1234567. Used ONLY to count/bound the
# data rows of a table (not to filter which cells get extracted).
CONTAINER_RE = re.compile(r"\b[A-Z]{4}\d{7}\b")

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# Spreadsheet extensions handled natively (per-engine), grouped by reader.
OPENPYXL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
XLS_EXTS = {".xls"}
XLSB_EXTS = {".xlsb"}
ODS_EXTS = {".ods"}
CSV_EXTS = {".csv", ".tsv", ".txt"}


SCRIPT_DIR = Path(__file__).resolve().parent
EXTRACTION_DIR = SCRIPT_DIR / "extraction"

IN_OUT_WORD_RE = re.compile(r"\b(?:in|out)\b", re.IGNORECASE)
GATE_IN_OUT_RE = re.compile(r"gate\s+(?:in|out)", re.IGNORECASE)
DAILY_MOVEMENT_RE = re.compile(r"daily\s+movement", re.IGNORECASE)


def _is_empty(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def _to_text(value):
    """Copy the cell's contents into the CSV verbatim. Everything is treated
    the same regardless of type — no stripping, no reformatting, no date/time
    handling, no locale changes. Only None becomes an empty cell."""
    if value is None:
        return ""
    return str(value)


def _trim_rows(rows):
    """Drop fully-empty leading/trailing rows and trailing empty columns,
    while preserving the tabular structure (blank cells kept in place)."""
    # Remove fully-empty rows from the top and bottom only.
    start, end = 0, len(rows)
    while start < end and all(_is_empty(c) for c in rows[start]):
        start += 1
    while end > start and all(_is_empty(c) for c in rows[end - 1]):
        end -= 1
    rows = rows[start:end]
    if not rows:
        return []

    # Determine the last column index that contains any non-empty cell.
    max_col = 0
    for row in rows:
        for idx, cell in enumerate(row):
            if not _is_empty(cell):
                max_col = max(max_col, idx + 1)

    return [row[:max_col] for row in rows]


def _write_csv(rows, out_path):
    rows = _trim_rows(rows)
    if not rows:
        return False
    with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        for row in rows:
            writer.writerow([_to_text(c) for c in row])
    return True


def _safe_name(name):
    keep = "".join(ch if ch.isalnum() or ch in " -_." else "_" for ch in str(name))
    keep = keep.strip().rstrip(".")           # Windows dislikes trailing dots
    return keep[:120] or "sheet"


def _cell_str(c):
    return "" if _is_empty(c) else str(c).strip()


def _is_target_sheet_name(sheet_name):
    sheet_name = sheet_name or ""
    # Keep month-based in/out reports out even if they otherwise match.
    if re.search(r"\bmonth\b", sheet_name, re.IGNORECASE) and IN_OUT_WORD_RE.search(sheet_name):
        return False
    if IN_OUT_WORD_RE.search(sheet_name):
        return True
    if GATE_IN_OUT_RE.search(sheet_name):
        return True
    return bool(DAILY_MOVEMENT_RE.search(sheet_name))


def _row_has_container(row):
    return any(CONTAINER_RE.search(_cell_str(c)) for c in row)


def _row_is_blank(row):
    return all(_is_empty(c) for c in row)


def _row_text_values(row):
    values = []
    for c in row:
        value = _cell_str(c)
        if value and value not in values:
            values.append(value)
    return values


def _is_text_cell(c):
    if _is_empty(c):
        return False
    if isinstance(c, (int, float, _dt.date, _dt.time)):
        return False
    s = str(c).strip()
    if not s or CONTAINER_RE.fullmatch(s):
        return False
    return True


def _is_header_row(row):
    if any(isinstance(c, (int, float, _dt.date, _dt.time)) for c in row if not _is_empty(c)):
        return False
    consec = max_consec = 0
    for c in row:
        if _is_text_cell(c):
            consec += 1
            if consec > max_consec:
                max_consec = consec
        else:
            consec = 0
    return max_consec >= 2


def detect_tables(grid):
    """Detect multiple tables within a single sheet grid.

    Returns a list of (title, table_rows). Logic:
      - A header row is any row with 2+ consecutive non-empty text-only cells.
      - Column span is derived from the header row alone (first to last
        non-empty cell in that row).
      - The body is container-number rows below the header, ending at the first
        fully blank row. Trailing non-blank non-container rows are kept.
      - If no container rows are found below the header, the header itself is
        the entire table.
      - The title is built from non-empty cells in up to 3 rows above the
        header, stopping if those rows are already consumed.
    """
    n = len(grid)
    tables = []
    consumed = set()

    for i in range(n):
        if i in consumed or not _is_header_row(grid[i]):
            continue

        hr = i

        # Column span from header row only.
        cmin = cmax = None
        for ci, c in enumerate(grid[hr]):
            if not _is_empty(c):
                cmin = ci if cmin is None else min(cmin, ci)
                cmax = ci if cmax is None else max(cmax, ci)
        if cmin is None:
            continue

        # Search for container rows below the header.
        r_end = hr
        found_container = False
        j = hr + 1
        while j < n:
            if _row_is_blank(grid[j]):
                break
            if _row_has_container(grid[j]):
                found_container = True
                r_end = j
            j += 1

        # Title: always inspect the 2 rows directly above the header; stop only
        # at a consumed row (owned by a prior table). Blank rows are skipped
        # silently; text rows (including expanded merged cells) are included.
        title_row_indices = []
        for step in range(1, 3):
            k = hr - step
            if k < 0 or k in consumed:
                break
            vals = _row_text_values(grid[k])
            if vals:
                title_row_indices.insert(0, k)
        title_cells = [value for k in title_row_indices for value in _row_text_values(grid[k])]
        title = " ".join(title_cells) if title_cells else None
        title_start = title_row_indices[0] if title_row_indices else hr

        if not found_container:
            # If the next non-blank row is also a header, this row is a section
            # title for that table. Leave it for the next header's title look-back.
            peek = hr + 1
            while peek < n and _row_is_blank(grid[peek]):
                peek += 1
            if peek < n and _is_header_row(grid[peek]):
                continue
            # No body: still emit the title rows plus the header as a table.
            table_rows = [list(grid[r][cmin:cmax + 1]) for r in range(title_start, hr + 1)]
            tables.append((title, table_rows))
            consumed.update(title_row_indices)
            consumed.add(hr)
            continue

        # Extend body downward past trailing non-blank non-container rows.
        while r_end + 1 < n and not _row_is_blank(grid[r_end + 1]):
            r_end += 1

        table_rows = [list(grid[r][cmin:cmax + 1]) for r in range(title_start, r_end + 1)]
        tables.append((title, table_rows))
        consumed.update(title_row_indices)
        consumed.update(range(hr, r_end + 1))

    return tables


def write_tables(sheets, out_dir, default_name):
    """sheets: iterable of (sheet_name, grid).

    Writes extraction/<file>/<sheet>/<title>.csv per detected table. Tables
    without a title fall back to the sheet name. Sheets with no detected table
    fall back to writing the whole sheet."""
    for sheet_name, grid in sheets:
        if not _is_target_sheet_name(sheet_name):
            print(f"  (sheet '{sheet_name}' skipped — not gate in/out or daily movement)")
            continue
        sheet_label = _safe_name(sheet_name or default_name)
        tables = detect_tables(grid)
        sheet_dir = out_dir / sheet_label

        if not tables:
            sheet_dir.mkdir(parents=True, exist_ok=True)
            if _write_csv(grid, sheet_dir / f"{sheet_label}.csv"):
                print(f"  -> {sheet_label}/{sheet_label}.csv (no tables detected)")
            else:
                print(f"  (sheet '{sheet_name}' empty, skipped)")
            continue

        sheet_dir.mkdir(parents=True, exist_ok=True)
        used = {}
        for title, rows in tables:
            base = _safe_name(title) if title else sheet_label
            count = used.get(base, 0)
            used[base] = count + 1
            name = base if count == 0 else f"{base}_{count + 1}"
            if _write_csv(rows, sheet_dir / f"{name}.csv"):
                print(f"  -> {sheet_label}/{name}.csv")


def read_xlsx(path):
    if load_workbook is None:
        print("  ! openpyxl is not installed. Run: pip install openpyxl")
        return None
    wb = load_workbook(path, data_only=True)
    sheets = []
    for sh in wb.worksheets:
        grid = [list(r) for r in sh.iter_rows(values_only=True)]
        sheets.append((sh.title, grid))
    wb.close()
    return sheets


def read_csv(path):
    with open(path, "r", newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
        rows = [row for row in csv.reader(fh, dialect)]
    return [(path.stem, rows)]


def read_xls(path):
    try:
        import xlrd
        from xlrd.xldate import xldate_as_datetime
    except ImportError:
        print("  ! xlrd is not installed. Run: pip install xlrd")
        return None
    book = xlrd.open_workbook(path)

    def _convert_row(sh, r):
        out = []
        for c in range(sh.ncols):
            ctype = sh.cell_type(r, c)
            value = sh.cell_value(r, c)
            # XL_CELL_DATE (3): a real date/time cell is stored as a serial
            # float by Excel. Render it back to readable text so it copies as
            # e.g. "22-11-2023 08:54" instead of "45824.54954". Cells that are
            # already plain text (type 1) are left untouched.
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xldate_as_datetime(value, book.datemode)
                except (ValueError, OverflowError):
                    out.append(value)
                    continue
                if (dt.hour, dt.minute, dt.second) == (0, 0, 0):
                    value = dt.strftime("%d-%m-%Y")
                else:
                    value = dt.strftime("%d-%m-%Y %H:%M")
            out.append(value)
        return out

    sheets = []
    for sh in book.sheets():
        grid = [_convert_row(sh, r) for r in range(sh.nrows)]
        sheets.append((sh.name, grid))
    return sheets


def read_with_pandas(path, engine=None, label=""):
    try:
        import pandas as pd
    except ImportError:
        print(f"  ! pandas is required for {label or path.suffix} files. "
              f"Run: pip install pandas{(' ' + engine) if engine else ''}")
        return None
    try:
        book = pd.read_excel(path, sheet_name=None, header=None,
                             engine=engine, dtype=object)
    except ImportError as e:
        print(f"  ! Missing engine for {path.suffix}: {e}")
        return None
    return [(title, df.where(df.notna(), None).values.tolist())
            for title, df in book.items()]


def process_file(path):
    path = Path(path).resolve()
    if not path.is_file():
        print(f"! Not found: {path}")
        return
    print(f"Processing {path.name}")
    ext = path.suffix.lower()
    if ext in CSV_EXTS:
        sheets = read_csv(path)
    elif ext in OPENPYXL_EXTS:
        sheets = read_xlsx(path)
    elif ext in XLS_EXTS:
        sheets = read_xls(path)
    elif ext in XLSB_EXTS:
        sheets = read_with_pandas(path, engine="pyxlsb", label=".xlsb")
    elif ext in ODS_EXTS:
        sheets = read_with_pandas(path, engine="odf", label=".ods")
    else:
        print(f"  (unknown type {ext}; attempting pandas fallback)")
        sheets = read_with_pandas(path, label=ext)

    if not sheets:
        return
    out_dir = EXTRACTION_DIR / _safe_name(path.stem)
    out_dir.mkdir(parents=True, exist_ok=True)
    write_tables(sheets, out_dir, default_name=path.stem)


def main():
    args = sys.argv[1:]
    if args:
        targets = args
    else:
        supported = (CSV_EXTS | OPENPYXL_EXTS | XLS_EXTS | XLSB_EXTS | ODS_EXTS)
        targets = [
            p for p in SCRIPT_DIR.iterdir()
            if p.suffix.lower() in supported
            and p.name != Path(__file__).name
            and not p.name.startswith("~$")
        ]
        if not targets:
            print("No files given and no supported spreadsheet files "
                  "found next to the script.")
            return
    for t in targets:
        process_file(t)


if __name__ == "__main__":
    main()
