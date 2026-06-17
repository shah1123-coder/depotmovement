#!/usr/bin/env python3
"""Extract IN movement records into one combined text report."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from countries import fallbacks_for_sheet, india, patterns_for_sheet
from database import write_gate_in_payloads
from movement import categorize_and_copy

DEPOT_REPORT_CODE = Path(__file__).resolve().parents[1] / "depot_report" / "code"
sys.path.insert(0, str(DEPOT_REPORT_CODE))
from db.icms_client import get_container_ids, get_container_info, plotin_records_exist


CSV_ROOT = Path(__file__).resolve().parents[1]
IN_DIR = CSV_ROOT / "results" / "IN"
OUTPUT_PATH = CSV_ROOT / "results" / "in.txt"
SEP = " | "


@dataclass(frozen=True)
class Record:
    container_no: str
    date: str
    time: str
    remark: str
    container_status: str = ""
    error_code: str = ""


def merged_cell_map(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int, Any]]:
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merge_map[(row, col)] = (min_col, min_row, max_col, max_row, value)
    return merge_map


def cell_value(sheet, row, col, merge_map) -> Any:
    merge = merge_map.get((row, col))
    if merge:
        start_col, start_row, _ec, _er, value = merge
        if row == start_row and col == start_col:
            return value
        return None
    return sheet.cell(row=row, column=col).value


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return " ".join(str(value).strip().split())


def find_remark_column(sheet, merge_map, markers: tuple[str, ...]) -> int | None:
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = display_value(cell_value(sheet, row, col, merge_map)).lower()
            if any(marker.lower() in value for marker in markers):
                return col
    return None


def first_container(values, pattern: re.Pattern[str] = india.IN_PATTERNS.container_number) -> str | None:
    for value in values:
        match = pattern.search(str(value or "").strip())
        if match:
            return match.group(0).upper()
    return None


def first_match(values, pattern: re.Pattern) -> Any:
    for value in values:
        if pattern.search(str(value or "").strip()):
            return value
    return None


def normalize_date(value, pattern=india.IN_PATTERNS.date,
                   formats=india.FALLBACKS.date_formats,
                   whitespace_pattern=india.FALLBACKS.whitespace_pattern) -> str:
    if isinstance(value, datetime) or isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    match = pattern.search(text)
    if not match:
        return text
    raw = match.group(0).replace(",", " ")
    raw = whitespace_pattern.sub(" ", raw).strip()
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def normalize_time(value, pattern=india.IN_PATTERNS.time, formats=india.FALLBACKS.time_formats) -> str:
    if isinstance(value, datetime) or isinstance(value, time):
        return value.strftime("%H:%M:%S.000")
    text = str(value or "").strip()
    match = pattern.search(text)
    if not match:
        return text
    raw_time = match.group(0)
    for fmt in formats:
        try:
            return datetime.strptime(raw_time, fmt).strftime("%H:%M:%S.000")
        except ValueError:
            continue
    hour, minute, second, _ = match.groups()
    return f"{int(hour):02d}:{minute}:{second or '00'}.000"


def extract_records(workbook_path: Path) -> list[Record]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    records: list[Record] = []
    try:
        for sheet in workbook.worksheets:
            patterns = patterns_for_sheet(sheet, "IN")
            fallbacks = fallbacks_for_sheet(sheet)
            merge_map = merged_cell_map(sheet)
            remark_col = find_remark_column(sheet, merge_map, fallbacks.in_remark_markers)
            for row in range(1, sheet.max_row + 1):
                raw_values = [cell_value(sheet, row, col, merge_map) for col in range(1, sheet.max_column + 1)]
                container_no = first_container(raw_values, patterns.container_number)
                if not container_no:
                    continue
                remark = ""
                if remark_col is not None:
                    remark = display_value(cell_value(sheet, row, remark_col, merge_map))
                records.append(Record(
                    container_no=container_no,
                    date=normalize_date(first_match(raw_values, patterns.date),
                                        patterns.date, fallbacks.date_formats, fallbacks.whitespace_pattern),
                    time=normalize_time(first_match(raw_values, patterns.time),
                                        patterns.time, fallbacks.time_formats),
                    remark=remark,
                ))
    finally:
        workbook.close()
    return validate_records(records)


def validate_records(records: list[Record]) -> list[Record]:
    container_ids = get_container_ids(record.container_no for record in records)
    container_info = get_container_info(record.container_no for record in records)
    duplicate_pairs = [
        (container_ids.get(record.container_no), record.date)
        for record in records
        if container_ids.get(record.container_no) is not None and record.date
    ]
    duplicates = plotin_records_exist(duplicate_pairs)
    validated: list[Record] = []

    for record in records:
        container_id = container_ids.get(record.container_no)
        info = container_info.get(record.container_no)
        status_id = info[0] if info is not None else None
        error_code = ""
        if container_id is None:
            error_code = "NO_CONTAINER_ID"
        elif record.date and (int(container_id), record.date) in duplicates:
            error_code = "DUPLICATE_RECORD"
        elif status_id not in (2, 7):
            error_code = "INVALID_CONTAINER_STATUS_ID"
        validated.append(Record(
            container_no=record.container_no,
            date=record.date,
            time=record.time,
            remark=record.remark,
            container_status="" if status_id is None else str(status_id),
            error_code=error_code,
        ))
    return validated


def table_lines(title: str, records: list[Record]) -> list[str]:
    headers = ["Container #", "Date", "Time", "Remark", "Cnt_status", "ErrorCode"]
    rows = [headers]
    rows.extend([[r.container_no, r.date, r.time, r.remark, r.container_status, r.error_code] for r in records])
    widths = [max(len(row[i]) for row in rows) for i in range(len(headers))]
    lines = [title]
    lines.append(SEP.join(v.ljust(widths[i]) for i, v in enumerate(rows[0])))
    lines.append(SEP.join("-" * w for w in widths))
    for row in rows[1:]:
        lines.append(SEP.join(v.ljust(widths[i]) for i, v in enumerate(row)))
    if not records:
        lines.append("(none)")
    return lines


def build_report(input_dir: Path | None = None, output_path: Path | None = None, synchronize: bool = True) -> Path:
    input_dir = input_dir or IN_DIR
    output_path = output_path or OUTPUT_PATH
    if synchronize and input_dir == IN_DIR:
        categorize_and_copy(only_direction="IN")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    records_by_depot: list[tuple[str, list[Record]]] = []
    for workbook_path in sorted(input_dir.glob("*.xlsx")):
        try:
            records = extract_records(workbook_path)
        except Exception as exc:
            print(f"IN extract failed for {workbook_path.name}: {exc}")
            records = []
        records_by_depot.append((workbook_path.stem, records))
        if lines:
            lines.append("")
        lines.extend(table_lines(workbook_path.stem, records))
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    write_gate_in_payloads(records_by_depot)
    return output_path


def main() -> None:
    print(build_report())


if __name__ == "__main__":
    main()
