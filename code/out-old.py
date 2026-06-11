#!/usr/bin/env python3
"""Extract OUT movement records into one combined text report."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from countries import fallbacks_for_sheet, india, patterns_for_sheet
from movement import categorize_and_copy

DEPOT_REPORT_CODE = Path(__file__).resolve().parents[2] / "depot report" / "code"
sys.path.insert(0, str(DEPOT_REPORT_CODE))
from db.icms_client import (
    get_booked_qty_by_booking_id,
    get_booking_ids_by_reference,
    get_container_ids,
    get_plotout_container_counts,
    plotout_records_exist,
)


CSV_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = CSV_ROOT / "results" / "OUT"
OUTPUT_PATH = CSV_ROOT / "results" / "out.txt"
SEP = " | "

@dataclass(frozen=True)
class Record:
    container_id: str
    booking_id: str
    seal_no: str
    plot_out_date: str
    plot_out_time: str
    transporter: str
    vehicle_no: str
    remarks: str
    error_code: str


def merged_cell_map(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int, Any]]:
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merge_map[(row, col)] = (min_col, min_row, max_col, max_row, value)
    return merge_map


def cell_value(
    sheet: Worksheet,
    row: int,
    col: int | None,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]],
) -> Any:
    if col is None:
        return None
    merge = merge_map.get((row, col))
    if merge:
        start_col, start_row, _end_col, _end_row, value = merge
        if row == start_row and col == start_col:
            return value
        return None
    return sheet.cell(row=row, column=col).value


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return " ".join(str(value).strip().split())


def find_header_column(
    sheet: Worksheet,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]],
    needle: str,
) -> int | None:
    needle = needle.lower()
    rightmost = None
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if needle in display_value(cell_value(sheet, row, col, merge_map)).lower():
                if rightmost is None or col > rightmost:
                    rightmost = col
    return rightmost


def find_header_column_for_markers(
    sheet: Worksheet,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]],
    markers: tuple[str, ...],
) -> int | None:
    matches = [
        column
        for marker in markers
        if (column := find_header_column(sheet, merge_map, marker)) is not None
    ]
    return max(matches) if matches else None


def first_container(
    values: list[Any], pattern: re.Pattern[str] = india.OUT_PATTERNS.container_number
) -> str | None:
    for value in values:
        match = pattern.search(str(value or "").strip())
        if match:
            return match.group(0).upper()
    return None


def last_match(values: list[Any], pattern: re.Pattern) -> Any:
    for value in reversed(values):
        if pattern.search(str(value or "").strip()):
            return value
    return None


def find_regex_column(
    sheet: Worksheet,
    pattern: re.Pattern[str],
) -> int | None:
    matches: list[int] = []
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if pattern.search(str(val or "").strip()):
                matches.append(col)
    return max(matches) if matches else None


def normalize_date(
    value: Any,
    pattern: re.Pattern[str] = india.OUT_PATTERNS.date,
    formats: tuple[str, ...] = india.FALLBACKS.date_formats,
    whitespace_pattern: re.Pattern[str] = india.FALLBACKS.whitespace_pattern,
) -> str:
    if isinstance(value, datetime) or isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value or "").strip()
    match = pattern.search(text)
    if not match:
        return text
    
    raw = match.group(0).replace(",", " ")
    raw = whitespace_pattern.sub(" ", raw).strip()
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def normalize_time(
    value: Any,
    pattern: re.Pattern[str] = india.OUT_PATTERNS.time,
    formats: tuple[str, ...] = india.FALLBACKS.time_formats,
) -> str:
    if isinstance(value, datetime) or isinstance(value, time):
        return value.strftime("%H:%M:%S.000")

    text = str(value or "").strip()
    match = pattern.search(text)
    if not match:
        return text
    
    raw_time = match.group(0)
    for fmt in formats:
        try:
            dt = datetime.strptime(raw_time, fmt)
            return dt.strftime("%H:%M:%S.000")
        except ValueError:
            continue

    hour, minute, second, _ = match.groups()
    return f"{int(hour):02d}:{minute}:{second or '00'}.000"


def extract_records(workbook_path: Path) -> list[Record]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    records: list[Record] = []
    try:
        for sheet in workbook.worksheets:
            patterns = patterns_for_sheet(sheet, "OUT")
            fallbacks = fallbacks_for_sheet(sheet)
            merge_map = merged_cell_map(sheet)
            
            # Booking ID: Regex -> Marker -> 6-digit
            booking_col = find_regex_column(sheet, patterns.booking_id)
            if not booking_col:
                booking_col = find_header_column_for_markers(
                    sheet, merge_map, fallbacks.out_booking_markers
                )
            if not booking_col:
                booking_col = find_regex_column(
                    sheet, fallbacks.out_booking_fallback
                )
            
            seal_col = find_header_column_for_markers(
                sheet, merge_map, fallbacks.out_seal_markers
            )
            transporter_col = find_header_column_for_markers(
                sheet, merge_map, fallbacks.out_transporter_markers
            )
            
            # Remarks: "remark" or Marker
            remarks_col = find_header_column_for_markers(
                sheet, merge_map, fallbacks.out_remark_markers
            )
            
            # Vehicle: Marker -> Regex fallback
            vehicle_col = find_header_column_for_markers(
                sheet, merge_map, fallbacks.out_vehicle_markers
            )
            
            if not vehicle_col:
                # Find representative rows with containers to identify other fixed columns
                representative_rows = []
                for row in range(1, sheet.max_row + 1):
                    vals = [cell_value(sheet, row, col, merge_map) for col in range(1, sheet.max_column + 1)]
                    if first_container(vals, patterns.container_number):
                        representative_rows.append(row)
                
                # Find vehicle column via representative rows
                for row in representative_rows:
                    for col in range(1, sheet.max_column + 1):
                        val = str(cell_value(sheet, row, col, merge_map) or "")
                        if patterns.vehicle_number.search(val):
                            vehicle_col = col
                            break
                    if vehicle_col: break

            for row in range(1, sheet.max_row + 1):
                raw_values = [cell_value(sheet, row, col, merge_map) for col in range(1, sheet.max_column + 1)]
                container_id = first_container(raw_values, patterns.container_number)
                if not container_id:
                    continue
                
                booking_id = display_value(cell_value(sheet, row, booking_col, merge_map))
                error_code = (
                    ""
                    if patterns.booking_id.search(booking_id)
                    else fallbacks.default_error_code
                )
                
                records.append(
                    Record(
                        container_id=container_id,
                        booking_id=booking_id,
                        seal_no=display_value(cell_value(sheet, row, seal_col, merge_map)),
                        plot_out_date=normalize_date(
                            last_match(raw_values, patterns.date),
                            patterns.date,
                            fallbacks.date_formats,
                            fallbacks.whitespace_pattern,
                        ),
                        plot_out_time=normalize_time(
                            last_match(raw_values, patterns.time),
                            patterns.time,
                            fallbacks.time_formats,
                        ),
                        transporter=display_value(cell_value(sheet, row, transporter_col, merge_map)),
                        vehicle_no=display_value(cell_value(sheet, row, vehicle_col, merge_map)),
                        remarks=display_value(cell_value(sheet, row, remarks_col, merge_map)),
                        error_code=error_code,
                    )
                )
    finally:
        workbook.close()
    return validate_records(records)


def validate_records(records: list[Record]) -> list[Record]:
    container_ids = get_container_ids(record.container_id for record in records)
    booking_refs = [record.booking_id for record in records if record.booking_id]
    booking_ids = get_booking_ids_by_reference(booking_refs)

    duplicate_keys: list[tuple[int, str, int] | None] = []
    duplicate_candidates: list[tuple[int, str, int]] = []
    for record in records:
        container_id = container_ids.get(record.container_id)
        booking_id = booking_ids.get(record.booking_id) if record.booking_id else None
        if container_id is not None and booking_id is not None and record.plot_out_date:
            key = (int(container_id), record.plot_out_date, int(booking_id))
            duplicate_keys.append(key)
            duplicate_candidates.append(key)
        else:
            duplicate_keys.append(None)
    duplicates = plotout_records_exist(duplicate_candidates)

    eligible_booking_ids: list[int | None] = []
    extracted_counts: dict[int, int] = {}
    for record, duplicate_key in zip(records, duplicate_keys):
        container_id = container_ids.get(record.container_id)
        booking_id = booking_ids.get(record.booking_id) if record.booking_id else None
        has_prior_error = bool(record.error_code)
        is_duplicate = duplicate_key is not None and duplicate_key in duplicates
        if (
            container_id is not None
            and booking_id is not None
            and not has_prior_error
            and not is_duplicate
        ):
            extracted_counts[booking_id] = extracted_counts.get(booking_id, 0) + 1
            eligible_booking_ids.append(booking_id)
        else:
            eligible_booking_ids.append(None)

    booked = get_booked_qty_by_booking_id(extracted_counts)
    existing = get_plotout_container_counts(extracted_counts)
    invalid_ecp = {
        booking_id
        for booking_id, extracted in extracted_counts.items()
        if extracted > booked.get(booking_id, 0) - existing.get(booking_id, 0)
    }

    validated: list[Record] = []
    for record, duplicate_key, booking_id in zip(
        records, duplicate_keys, eligible_booking_ids
    ):
        error_code = record.error_code
        if container_ids.get(record.container_id) is None:
            error_code = "NO_CONTAINER_ID"
        elif not error_code and duplicate_key is not None and duplicate_key in duplicates:
            error_code = "DUPLICATE_RECORD"
        elif not error_code and booking_id in invalid_ecp:
            error_code = "INVALID_ECP_COUNT"

        validated.append(
            Record(
                container_id=record.container_id,
                booking_id=record.booking_id,
                seal_no=record.seal_no,
                plot_out_date=record.plot_out_date,
                plot_out_time=record.plot_out_time,
                transporter=record.transporter,
                vehicle_no=record.vehicle_no,
                remarks=record.remarks,
                error_code=error_code,
            )
        )
    return validated


def table_lines(title: str, records: list[Record]) -> list[str]:
    headers = [
        "ContainerID", "BookingId", "SealNo", "PlotOutDate", "PlotOutTime",
        "Transporter", "VehicleNo", "Remarks", "ErrorCode",
    ]
    rows = [headers]
    rows.extend([
        [
            record.container_id,
            record.booking_id,
            record.seal_no,
            record.plot_out_date,
            record.plot_out_time,
            record.transporter,
            record.vehicle_no,
            record.remarks,
            record.error_code,
        ]
        for record in records
    ])
    widths = [max(len(row[index]) for row in rows) for index in range(len(headers))]

    lines = [title]
    lines.append(SEP.join(value.ljust(widths[index]) for index, value in enumerate(rows[0])))
    lines.append(SEP.join("-" * width for width in widths))
    for row in rows[1:]:
        lines.append(SEP.join(value.ljust(widths[index]) for index, value in enumerate(row)))
    if not records:
        lines.append("(none)")
    return lines


def build_report(
    input_dir: Path = OUT_DIR,
    output_path: Path = OUTPUT_PATH,
    synchronize: bool = True,
) -> Path:
    if synchronize and input_dir == OUT_DIR:
        categorize_and_copy(only_direction="OUT")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    for workbook_path in sorted(input_dir.glob("*.xlsx")):
        if lines:
            lines.append("")
        lines.extend(table_lines(workbook_path.stem, extract_records(workbook_path)))
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


def main() -> None:
    print(build_report())


if __name__ == "__main__":
    main()
