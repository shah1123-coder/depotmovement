#!/usr/bin/env python3
"""
Detect spreadsheet headers.

A header is a continuous run on the same row containing at least 3 logical cells.
Each logical cell must contain plain text. A horizontal merged range is treated as
one logical header cell, and its text is extracted from the merged range's
top-left cell.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from converter import ensure_xlsx
from countries import sheet_selection_for_name


CONTAINER_NUMBER = re.compile(r"\b[A-Z]{3}[UJZ]\d{7}\b", re.IGNORECASE)
CSV_ROOT = Path(__file__).resolve().parents[1]
PROCESSED_DIR = CSV_ROOT / "processed"
EXTRACTION_DIR = CSV_ROOT / "extraction"


@dataclass(frozen=True)
class LogicalCell:
    row: int
    end_row: int
    start_col: int
    end_col: int
    value: Any
    is_horizontal_merge: bool
    is_vertical_merge: bool


@dataclass(frozen=True)
class ExtractedCell:
    row: int
    end_row: int
    start_col: int
    end_col: int
    value: str
    is_horizontal_merge: bool
    is_vertical_merge: bool


@dataclass(frozen=True)
class TableTitle:
    row: int
    cells: list[ExtractedCell]
    text: str


@dataclass(frozen=True)
class HeaderRun:
    sheet: str
    row: int
    start_col: int
    end_col: int
    logical_cell_count: int
    values: list[str]
    cells: list[ExtractedCell]
    title: TableTitle | None


@dataclass(frozen=True)
class TableExtract:
    index: int
    header: HeaderRun
    data_rows: list[list[ExtractedCell]]
    output: str | None = None


def is_plain_text(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    return not text.startswith("=")


def merged_cell_map(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int, Any]]:
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merge_map[(row, col)] = (min_col, min_row, max_col, max_row, value)
    return merge_map


def iter_logical_row_cells(
    sheet: Worksheet, row: int, merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]]
) -> list[LogicalCell]:
    cells: list[LogicalCell] = []
    col = 1
    while col <= sheet.max_column:
        merge = merge_map.get((row, col))
        if merge:
            start_col, start_row, end_col, end_row, value = merge
            if col == start_col:
                cells.append(
                    LogicalCell(
                        row=row,
                        end_row=end_row,
                        start_col=start_col,
                        end_col=end_col,
                        value=value,
                        is_horizontal_merge=end_col > start_col,
                        is_vertical_merge=end_row > start_row,
                    )
                )
            col = end_col + 1
            continue

        cells.append(
            LogicalCell(
                row=row,
                end_row=row,
                start_col=col,
                end_col=col,
                value=sheet.cell(row=row, column=col).value,
                is_horizontal_merge=False,
                is_vertical_merge=False,
            )
        )
        col += 1
    return cells


def detect_headers_in_sheet(sheet: Worksheet, min_cells: int = 3) -> list[HeaderRun]:
    merge_map = merged_cell_map(sheet)
    headers: list[HeaderRun] = []
    row = 1
    while row <= sheet.max_row:
        runs = candidate_header_runs(sheet, row, merge_map, min_cells)
        if not runs:
            row += 1
            continue
        for run in runs:
            headers.append(header_from_run(sheet.title, run, find_title(sheet, row, merge_map)))
        row += 1
        while row <= sheet.max_row and candidate_header_runs(sheet, row, merge_map, min_cells):
            row += 1
    return headers


def candidate_header_runs(
    sheet: Worksheet,
    row: int,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]],
    min_cells: int,
) -> list[list[LogicalCell]]:
    runs: list[list[LogicalCell]] = []
    run: list[LogicalCell] = []
    for cell in iter_logical_row_cells(sheet, row, merge_map):
        if is_plain_text(cell.value):
            run.append(cell)
            continue
        if len(run) >= min_cells:
            runs.append(run)
        run = []
    if len(run) >= min_cells:
        runs.append(run)
    return runs


def has_content(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def find_title(
    sheet: Worksheet, header_row: int, merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]]
) -> TableTitle | None:
    for row in range(header_row - 1, max(header_row - 3, 0), -1):
        logical_cells = iter_logical_row_cells(sheet, row, merge_map)
        content_cells = [
            ExtractedCell(
                row=cell.row,
                end_row=cell.end_row,
                start_col=cell.start_col,
                end_col=cell.end_col,
                value=str(cell.value).strip(),
                is_horizontal_merge=cell.is_horizontal_merge,
                is_vertical_merge=cell.is_vertical_merge,
            )
            for cell in logical_cells
            if has_content(cell.value)
        ]
        if content_cells:
            return TableTitle(
                row=row,
                cells=content_cells,
                text=" ".join(cell.value for cell in content_cells),
            )
    return None


def header_from_run(sheet_name: str, run: list[LogicalCell], title: TableTitle | None) -> HeaderRun:
    cells = [
        ExtractedCell(
            row=cell.row,
            end_row=cell.end_row,
            start_col=cell.start_col,
            end_col=cell.end_col,
            value=str(cell.value).strip(),
            is_horizontal_merge=cell.is_horizontal_merge,
            is_vertical_merge=cell.is_vertical_merge,
        )
        for cell in run
    ]
    return HeaderRun(
        sheet=sheet_name,
        row=run[0].row,
        start_col=run[0].start_col,
        end_col=run[-1].end_col,
        logical_cell_count=len(run),
        values=[cell.value for cell in cells],
        cells=cells,
        title=title,
    )


def detect_headers(workbook_path: Path, min_cells: int = 3) -> list[HeaderRun]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    headers: list[HeaderRun] = []
    for sheet in workbook.worksheets:
        if not should_process_sheet(sheet.title):
            continue
        headers.extend(detect_headers_in_sheet(sheet, min_cells=min_cells))
    return headers


def detect_tables(workbook_path: Path, min_cells: int = 3) -> list[TableExtract]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    tables: list[TableExtract] = []
    for sheet in workbook.worksheets:
        if not should_process_sheet(sheet.title):
            continue
        merge_map = merged_cell_map(sheet)
        headers = detect_headers_in_sheet(sheet, min_cells=min_cells)
        for header in headers:
            tables.append(
                TableExtract(
                    index=len(tables) + 1,
                    header=header,
                    data_rows=extract_data_rows(sheet, header, merge_map),
                )
            )
    return tables


def prepare_input_workbook(workbook_path: Path) -> Path:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    source = workbook_path.resolve()
    source = ensure_xlsx(source, PROCESSED_DIR)
    target = PROCESSED_DIR / source.name
    if source == target.resolve():
        return target
    shutil.copy2(source, target)
    return target


def extraction_root_for_workbook(workbook_path: Path) -> Path:
    workbook_folder = sanitize_path_part(workbook_path.stem)
    root = EXTRACTION_DIR / workbook_folder
    root.mkdir(parents=True, exist_ok=True)
    return root


def extract_data_rows(
    sheet: Worksheet,
    header: HeaderRun,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int, Any]],
) -> list[list[ExtractedCell]]:
    rows: list[list[ExtractedCell]] = []
    for row in range(header.row + 1, sheet.max_row + 1):
        logical_cells = iter_logical_row_cells(sheet, row, merge_map)
        if not row_has_content(logical_cells):
            break
        extracted = extract_cells_for_columns(logical_cells, header.start_col, header.end_col)
        extracted.append(
            ExtractedCell(
                row=row,
                end_row=row,
                start_col=header.end_col + 1,
                end_col=header.end_col + 1,
                value="" if row_has_container_number(logical_cells) else "INVALID_CONTAINER_NUMBER",
                is_horizontal_merge=False,
                is_vertical_merge=False,
            )
        )
        rows.append(extracted)
    return rows


def row_has_content(cells: list[LogicalCell]) -> bool:
    return any(has_content(cell.value) for cell in cells)


def row_has_container_number(cells: list[LogicalCell]) -> bool:
    return any(CONTAINER_NUMBER.search(str(cell.value or "")) for cell in cells)


def extract_cells_for_columns(
    cells: list[LogicalCell], start_col: int, end_col: int
) -> list[ExtractedCell]:
    extracted: list[ExtractedCell] = []
    for cell in cells:
        if cell.end_col < start_col or cell.start_col > end_col:
            continue
        extracted.append(
            ExtractedCell(
                row=cell.row,
                end_row=cell.end_row,
                start_col=max(cell.start_col, start_col),
                end_col=min(cell.end_col, end_col),
                value="" if cell.value is None else str(cell.value).strip(),
                is_horizontal_merge=cell.end_col > cell.start_col,
                is_vertical_merge=cell.is_vertical_merge,
            )
        )
    return extracted


def should_process_sheet(sheet_name: str) -> bool:
    selection = sheet_selection_for_name(sheet_name)
    normalized = sheet_name.upper()
    if any(part in normalized for part in selection.excluded_name_parts):
        return False
    if normalized in selection.always_process_names:
        return True
    if normalized in selection.direct_in_names or normalized in selection.direct_out_names:
        return True

    has_in = (
        bool(selection.gate_in.search(sheet_name))
        if selection.in_language_marker and selection.in_language_marker in sheet_name
        else bool(selection.standalone_in.search(sheet_name))
    )
    has_out = (
        bool(selection.gate_out.search(sheet_name))
        if selection.out_language_marker and selection.out_language_marker in sheet_name
        else bool(selection.standalone_out.search(sheet_name))
    )

    if bool(selection.recognized_names.search(sheet_name)):
        return not (has_in and has_out)
    return (has_in or has_out) and not (has_in and has_out)


def write_table_workbooks(tables: list[TableExtract], output_path: Path) -> list[TableExtract]:
    saved_tables: list[TableExtract] = []
    for table in tables:
        sheet_dir = output_path / sanitize_path_part(table.header.sheet)
        sheet_dir.mkdir(parents=True, exist_ok=True)
        table_name = table_filename_stem(table)
        table_path = sheet_dir / f"{table_name}.xlsx"
        saved_path = write_table_workbook(table, table_path)
        saved_tables.append(
            TableExtract(
                index=table.index,
                header=table.header,
                data_rows=table.data_rows,
                output=str(saved_path),
            )
        )
    return saved_tables


def table_filename_stem(table: TableExtract) -> str:
    name = table.header.title.text if table.header.title else table.header.sheet
    return sanitize_path_part(name)[:80] or "table"


def sanitize_path_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(".")
    return cleaned or "untitled"


def write_table_workbook(table: TableExtract, output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table"
    col_offset = table.header.start_col - 1

    if table.header.title:
        write_extracted_cells(sheet, 1, table.header.title.cells, col_offset=col_offset)
        header_output_row = 1 + (table.header.row - table.header.title.row)
    else:
        header_output_row = 1

    write_extracted_cells(sheet, header_output_row, table.header.cells, col_offset=col_offset)
    sheet.cell(
        row=header_output_row,
        column=table.header.end_col - col_offset + 1,
        value="Error",
    )

    current_output_row = header_output_row + 1
    for data_row in table.data_rows:
        write_extracted_cells(sheet, current_output_row, data_row, col_offset=col_offset)
        current_output_row += 1

    saved_path = available_output_path(output_path)
    workbook.save(saved_path)
    return saved_path


def available_output_path(output_path: Path) -> Path:
    if not output_path.exists():
        return output_path
    try:
        with output_path.open("a+b"):
            return output_path
    except PermissionError:
        pass
    for index in range(1, 1000):
        fallback = output_path.with_name(f"{output_path.stem}_{index}{output_path.suffix}")
        if not fallback.exists():
            return fallback
        try:
            with fallback.open("a+b"):
                return fallback
        except PermissionError:
            continue
    raise PermissionError(f"Could not find an unlocked output path near {output_path}")


def write_extracted_cells(
    sheet: Worksheet, output_row: int, cells: list[ExtractedCell], col_offset: int = 0
) -> None:
    for cell in cells:
        output_start_col = max(1, cell.start_col - col_offset)
        output_end_col = max(output_start_col, cell.end_col - col_offset)
        sheet.cell(row=output_row, column=output_start_col, value=cell.value)
        if output_end_col > output_start_col:
            start = get_column_letter(output_start_col)
            end = get_column_letter(output_end_col)
            sheet.merge_cells(f"{start}{output_row}:{end}{output_row}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Detect text headers in .xlsx sheets.")
    parser.add_argument("workbook", type=Path, help="Path to an Excel workbook")
    parser.add_argument("--output", type=Path, help="Optional extraction root.")
    parser.add_argument("--min-cells", type=int, default=3,
                        help="Minimum logical text cells needed for a header run. Default: 3")
    args = parser.parse_args()

    processed_workbook = prepare_input_workbook(args.workbook)
    output_root = args.output or extraction_root_for_workbook(processed_workbook)
    output_root.mkdir(parents=True, exist_ok=True)
    tables = detect_tables(processed_workbook, min_cells=args.min_cells)
    saved_tables = write_table_workbooks(tables, output_root)

    payload = {
        "processed_workbook": str(processed_workbook),
        "extraction_root": str(output_root),
        "has_tables": bool(saved_tables),
        "outputs": [table.output for table in saved_tables if table.output],
        "tables": [asdict(table) for table in saved_tables],
    }
    print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
