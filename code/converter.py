#!/usr/bin/env python3
"""Convert legacy Excel BIFF workbooks to .xlsx."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook


LEGACY_EXCEL_SUFFIXES = {".xls", ".xlt"}
OPENXML_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_EXCEL_SUFFIXES = LEGACY_EXCEL_SUFFIXES | OPENXML_EXCEL_SUFFIXES


def converted_cell_value(cell: Any, datemode: int) -> Any:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, "#ERROR!")
    return cell.value


def convert_legacy_workbook(source: Path, target: Path) -> Path:
    source = source.resolve()
    target = target.resolve()
    if source.suffix.lower() not in LEGACY_EXCEL_SUFFIXES:
        raise ValueError(f"Unsupported legacy Excel format: {source.suffix}")

    legacy_book = xlrd.open_workbook(source, formatting_info=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for legacy_sheet in legacy_book.sheets():
        sheet = workbook.create_sheet(legacy_sheet.name[:31])
        for row in range(legacy_sheet.nrows):
            for col in range(legacy_sheet.ncols):
                cell = legacy_sheet.cell(row, col)
                sheet.cell(
                    row=row + 1,
                    column=col + 1,
                    value=converted_cell_value(cell, legacy_book.datemode),
                )

        for row_start, row_end, col_start, col_end in legacy_sheet.merged_cells:
            sheet.merge_cells(
                start_row=row_start + 1,
                end_row=row_end,
                start_column=col_start + 1,
                end_column=col_end,
            )

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def ensure_xlsx(source: Path, target_dir: Path) -> Path:
    suffix = source.suffix.lower()
    if suffix in OPENXML_EXCEL_SUFFIXES:
        return source
    if suffix in LEGACY_EXCEL_SUFFIXES:
        return convert_legacy_workbook(source, target_dir / f"{source.stem}.xlsx")
    supported = ", ".join(sorted(SUPPORTED_EXCEL_SUFFIXES))
    raise ValueError(f"Unsupported Excel format '{source.suffix}'. Supported: {supported}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert a legacy Excel workbook to .xlsx.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.workbook.with_suffix(".xlsx")
    print(convert_legacy_workbook(args.workbook, output))


if __name__ == "__main__":
    main()
