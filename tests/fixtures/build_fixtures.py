"""Generate small sample workbooks used by the unit/e2e tests."""

from pathlib import Path

from openpyxl import Workbook

FIX = Path(__file__).resolve().parent


def india_gate_in() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "GATE IN"
    ws["A1"] = "Daily Gate In Report"
    ws.append([])  # spacer handled below; overwrite rows explicitly
    ws["A2"] = "Container No"
    ws["B2"] = "Date"
    ws["C2"] = "Time"
    ws["D2"] = "Remark"
    ws["A3"] = "TCNU1234567"
    ws["B3"] = "2026-06-18"
    ws["C3"] = "09:30"
    ws["D3"] = "ok"
    ws["A4"] = "MEDU7654321"
    ws["B4"] = "2026-06-18"
    ws["C4"] = "10:15"
    ws["D4"] = "fine"
    path = FIX / "india_gate_in.xlsx"
    wb.save(path)
    return path


def india_gate_out() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "GATE OUT"
    ws["A1"] = "Container No"
    ws["B1"] = "Booking"
    ws["C1"] = "Seal"
    ws["D1"] = "Date"
    ws["E1"] = "Time"
    ws["A2"] = "TCNU1234567"
    ws["B2"] = "ABC/DEF/123456"
    ws["C2"] = "SEAL001"
    ws["D2"] = "2026-06-18"
    ws["E2"] = "14:30"
    path = FIX / "india_gate_out.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    print(india_gate_in())
    print(india_gate_out())
